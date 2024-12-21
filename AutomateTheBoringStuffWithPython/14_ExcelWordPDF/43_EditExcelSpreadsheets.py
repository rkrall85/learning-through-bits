

import openpyxl

wb = openpyxl.Workbook()

sheet = wb.get_sheet_by_name('Sheet') #create sheet var
#assign new values
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'

import os

os.chdir('C:\\Users\\rkrall\\github\\AutomateTheBoringStuffWithPython\\14_ExcelWordPDF')

sheet2 = wb.create_sheet() #create new sheet

sheet2.title = 'My new Sheet Name'

#print(wb.get_sheet_names())
wb.create_sheet(index=0, title='My Other Sheet') #put the sheet in certain order

wb.save('43_zExcelSample.xlsx') #will overwrite what is there.
