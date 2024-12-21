



import openpyxl
import os
os.chdir('C:\\Users\\rkrall\\github\\AutomateTheBoringStuffWithPython\\14_ExcelWordPDF')


workbook = openpyxl.load_workbook('42_zExcelSample.xlsx')
#print(type(workbook))


sheet = workbook.get_sheet_by_name('Sheet1')

#print(sheet)
#print(workbook.get_sheet_names()) #output all sheet names


cell = sheet['A1']
#print(cell.value) #output cell value
#print(str(sheet['A1'].value)) #print in a string
#print(sheet.cell(row=1, column=1).value) #output in cell method

for i in range(1,8):
    print(i, sheet.cell(row=i, column=2).value)
